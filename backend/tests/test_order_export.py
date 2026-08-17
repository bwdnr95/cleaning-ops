import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook


def test_admin_order_export_xlsx_preserves_requested_order(
    client: TestClient,
    seed_admin_token: str,
) -> None:
    headers = {"Authorization": f"Bearer {seed_admin_token}"}
    create = client.post(
        "/api/admin/orders/groups",
        headers=headers,
        json={
            "customer_name": "Export Customer",
            "customer_phone": "010-1111-2222",
            "customer_address": "Seoul",
            "customer_address_detail": "Unit 101",
            "source_channel": "naver",
            "notes": "internal memo",
            "lines": [
                {
                    "received_date": "2026-06-01",
                    "visit_dates": ["2026-06-02", "2026-06-04", "2026-06-08"],
                    "requested_time": "10:00",
                    "service_name": "Export line A",
                    "size_or_quantity": "12.0sqm",
                    "total_amount": 100000,
                    "discount_amount": 5000,
                    "deposit_amount": 10000,
                    "balance_amount": 90000,
                    "onsite_extra_amount": 0,
                    "vat_type": "included",
                    "payment_status": "deposit_paid",
                    "payment_memo": "payment memo A",
                    "evidence_memo": "evidence memo A",
                    "partner_payment_amount": 40000,
                    "partner_payment_status": "unpaid",
                },
                {
                    "received_date": "2026-06-01",
                    "scheduled_date": "2026-06-03",
                    "requested_time": "11:00",
                    "service_name": "Export line B",
                    "size_or_quantity": "3.50ea",
                    "total_amount": 120000,
                    "partner_payment_amount": 60000,
                    "partner_payment_status": "paid",
                },
            ],
        },
    )
    assert create.status_code == 201, create.text
    lines_by_service = {line["service_name"]: line for line in create.json()["lines"]}
    line_a = lines_by_service["Export line A"]
    line_b = lines_by_service["Export line B"]

    response = client.post(
        "/api/admin/orders/export",
        headers=headers,
        json={"order_ids": [line_b["id"], line_a["id"]]},
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="orders.xlsx"' in response.headers.get("content-disposition", "")

    workbook = load_workbook(io.BytesIO(response.content), read_only=False)
    try:
        rows = list(workbook["orders"].iter_rows(values_only=True))
    finally:
        workbook.close()

    header = list(rows[0])
    body = rows[1:]
    assert "입금메모" in header
    assert "증빙메모" in header
    assert "이윤" in header
    assert "내부메모" in header
    assert "전체 방문일" in header

    order_id_index = header.index("주문ID")
    service_index = header.index("상품명")
    quantity_index = header.index("수량")
    profit_index = header.index("이윤")
    notes_index = header.index("내부메모")
    visit_dates_index = header.index("전체 방문일")

    assert [row[order_id_index] for row in body] == [line_b["id"], line_a["id"]]
    assert [row[service_index] for row in body] == ["Export line B", "Export line A"]
    assert [row[quantity_index] for row in body] == ["3.5ea", "12sqm"]
    assert [row[profit_index] for row in body] == [60000, 60000]
    assert [row[notes_index] for row in body] == ["internal memo", "internal memo"]
    assert body[1][visit_dates_index] == "2026-06-02, 2026-06-04, 2026-06-08"

    worksheet = workbook["orders"]
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:AJ3"
    assert worksheet["A1"].font.bold is True
    assert worksheet["A1"].font.color.rgb == "00FFFFFF"
    assert worksheet["A1"].fill.fgColor.rgb == "001F2937"
    assert worksheet.column_dimensions["J"].width == 34
    assert worksheet.column_dimensions["O"].width == 28
    assert worksheet["J2"].alignment.wrap_text is True
    assert worksheet["Q2"].number_format == "#,##0"


def test_admin_order_export_empty_order_ids_returns_400(
    client: TestClient,
    seed_admin_token: str,
) -> None:
    response = client.post(
        "/api/admin/orders/export",
        headers={"Authorization": f"Bearer {seed_admin_token}"},
        json={"order_ids": []},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "order_ids_required"


def test_admin_order_export_deleted_order_fails_explicitly(
    client: TestClient,
    seed_admin_token: str,
    seed_order_id: str,
) -> None:
    headers = {"Authorization": f"Bearer {seed_admin_token}"}
    delete_response = client.delete(f"/api/admin/orders/{seed_order_id}", headers=headers)
    assert delete_response.status_code == 204

    response = client.post(
        "/api/admin/orders/export",
        headers=headers,
        json={"order_ids": [seed_order_id]},
    )

    assert response.status_code == 404
    assert response.json()["detail"] == {
        "code": "orders_not_found",
        "order_ids": [seed_order_id],
    }


def test_admin_order_export_requires_admin(
    client: TestClient,
    seed_partner_token: str,
    seed_order_id: str,
) -> None:
    response = client.post(
        "/api/admin/orders/export",
        headers={"Authorization": f"Bearer {seed_partner_token}"},
        json={"order_ids": [seed_order_id]},
    )

    assert response.status_code == 403
