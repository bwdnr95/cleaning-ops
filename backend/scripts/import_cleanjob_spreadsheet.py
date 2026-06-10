from __future__ import annotations

import argparse
import hashlib
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from secrets import token_urlsafe
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

BACKEND_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_ROOT))

from app.core.config import settings
from app.db.session import SessionLocal
from app.domain.constants import OrderStatus, TimelineEventType, UserRole, VatType
from app.domain.partner_category import DEFAULT_PARTNER_CATEGORIES, infer_partner_category_id
from app.domain.payment_status import PaymentStatus, PartnerPaymentStatus
from app.domain.partner_vat import gross_up_partner_vat_amount, should_gross_up_partner_vat_amount
from app.domain.phone import normalize_phone
from app.models import Order, OrderGroup, OrderTimeline, Partner, PartnerCategory, User


REPO_ROOT = Path(__file__).resolve().parents[2]
IMPORT_ID = "legacy260526"
DEFAULT_HEADER_ROW = 5
DEFAULT_DATA_ROW = 7
PHONE_PATTERN = re.compile(r"0\d{1,2}[-\s]?\d{3,4}[-\s]?\d{4}")
MONEY_PATTERN = re.compile(r"[^0-9.\-]")
SETTLED_DATE_PATTERN = re.compile(r"(\d{1,2})\s*/\s*(\d{1,2})")


@dataclass(frozen=True)
class SheetSpec:
    slug: str
    is_recurring: bool
    status: int = 1
    received_date: int = 2
    scheduled_date: int = 3
    requested_time: int = 4
    team_name: int = 5
    service_name: int = 6
    address: int = 7
    size_or_quantity: int = 8
    service_detail: int = 9
    source_channel: int = 10
    customer_name: int = 11
    customer_phone: int = 12
    total_amount: int = 13
    deposit_amount: int | None = 14
    balance_amount: int | None = 15
    onsite_extra_amount: int | None = 16
    vat_type: int | None = 17
    payment_status: int | None = 18
    payment_memo: int | None = 19
    evidence_memo: int | None = 20
    partner_name: int | None = 21
    partner_payment_amount: int | None = 22
    partner_payment_status: int | None = 23
    billing_basis: int | None = None


SHEET_SPECS: dict[str, SheetSpec] = {
    "SERVICE ORDERS": SheetSpec(slug="svc", is_recurring=False),
    "유니클린텍": SheetSpec(slug="uni", is_recurring=False),
    "정기청소": SheetSpec(
        slug="reg",
        is_recurring=True,
        address=12,
        size_or_quantity=7,
        service_detail=8,
        source_channel=9,
        customer_name=10,
        customer_phone=11,
        total_amount=13,
        deposit_amount=None,
        balance_amount=15,
        onsite_extra_amount=None,
        vat_type=16,
        payment_status=17,
        payment_memo=18,
        evidence_memo=19,
        partner_name=20,
        partner_payment_amount=21,
        partner_payment_status=22,
        billing_basis=14,
    ),
}

DIRECT_STATUS_MAP = {status.value: status for status in OrderStatus}
LEGACY_STATUS_MAP = {
    "예약완료": OrderStatus.SCHEDULE_CONFIRMED,
    "계약완료": OrderStatus.SCHEDULE_CONFIRMED,
}


@dataclass(frozen=True)
class ImportIssue:
    sheet: str
    row_index: int
    severity: str
    reason: str


@dataclass(frozen=True)
class ParsedOrder:
    sheet: str
    sheet_slug: str
    row_index: int
    order_id: str
    preferred_group_id: str
    status: OrderStatus
    received_date: date
    scheduled_date: date | None
    requested_time: str | None
    team_name: str | None
    service_name: str
    size_or_quantity: str | None
    service_detail: str | None
    special_request: str | None
    source_channel: str | None
    customer_name: str
    customer_phone: str
    customer_address: str
    total_amount: Decimal | None
    discount_amount: Decimal
    deposit_amount: Decimal | None
    balance_amount: Decimal | None
    onsite_extra_amount: Decimal | None
    vat_type: str | None
    payment_status: str | None
    payment_memo: str | None
    evidence_memo: str | None
    partner_name: str | None
    partner_payment_amount: Decimal | None
    partner_payment_status: str | None
    partner_settled_at: datetime | None
    customer_visible_payment: bool = False


@dataclass
class ParseResult:
    rows: list[ParsedOrder] = field(default_factory=list)
    issues: list[ImportIssue] = field(default_factory=list)
    skipped_rows: int = 0

    @property
    def errors(self) -> list[ImportIssue]:
        return [issue for issue in self.issues if issue.severity == "error"]

    @property
    def warnings(self) -> list[ImportIssue]:
        return [issue for issue in self.issues if issue.severity == "warning"]


@dataclass
class ApplyStats:
    groups_created: int = 0
    groups_updated: int = 0
    orders_created: int = 0
    orders_updated: int = 0
    partners_created: int = 0
    timeline_events_created: int = 0


@dataclass(frozen=True)
class DatabasePlan:
    orders_to_create: int
    orders_to_update: int
    partners_to_create: int


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Dry-run or import the 260526 CleanJob legacy workbook.",
    )
    parser.add_argument("--path", type=Path, default=None)
    parser.add_argument("--sheet", choices=sorted(SHEET_SPECS), default=None)
    parser.add_argument("--start-date", type=parse_date_arg, default=None)
    parser.add_argument("--apply", action="store_true", help="Write parsed rows to the database.")
    parser.add_argument(
        "--allow-errors",
        action="store_true",
        help="Import valid rows even when parse errors are present.",
    )
    args = parser.parse_args()

    path = args.path or find_default_spreadsheet()
    result = read_rows(path, sheet_filter=args.sheet, start_date=args.start_date)

    with SessionLocal() as db:
        plan = build_database_plan(db, result.rows)
        print_summary(path=path, result=result, plan=plan)

        if not args.apply:
            print("\nDRY RUN: no database changes were written. Pass --apply to import.")
            return

        if result.errors and not args.allow_errors:
            print("\nImport aborted: parse errors exist. Re-run with --allow-errors to skip them.")
            raise SystemExit(1)

        stats = apply_rows(db, result.rows)
        db.commit()

    print("\nIMPORT COMPLETE")
    print(
        "groups created={0.groups_created}, updated={0.groups_updated}; "
        "orders created={0.orders_created}, updated={0.orders_updated}; "
        "partners created={0.partners_created}; timeline events created={0.timeline_events_created}".format(
            stats
        )
    )


def find_default_spreadsheet() -> Path:
    matches = sorted((REPO_ROOT / ".claude").glob("260526_*.xlsx"))
    if matches:
        return matches[0]
    raise FileNotFoundError("No .claude/260526_*.xlsx workbook was found.")


def read_rows(path: Path, *, sheet_filter: str | None, start_date: date | None) -> ParseResult:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result = ParseResult()
    try:
        for worksheet in workbook.worksheets:
            spec = SHEET_SPECS.get(worksheet.title)
            if spec is None or (sheet_filter and worksheet.title != sheet_filter):
                continue
            for row_index, values in enumerate(
                worksheet.iter_rows(min_row=DEFAULT_DATA_ROW, values_only=True),
                start=DEFAULT_DATA_ROW,
            ):
                if is_blank_legacy_row(values, spec):
                    result.skipped_rows += 1
                    continue
                try:
                    parsed, warnings = parse_row(
                        worksheet.title,
                        spec,
                        row_index,
                        values,
                    )
                except ValueError as exc:
                    result.issues.append(
                        ImportIssue(worksheet.title, row_index, "error", str(exc))
                    )
                    continue

                comparison_date = parsed.scheduled_date or parsed.received_date
                if start_date and comparison_date < start_date:
                    result.skipped_rows += 1
                    continue

                result.rows.append(parsed)
                result.issues.extend(warnings)
    finally:
        workbook.close()
    return result


def parse_row(
    sheet: str,
    spec: SheetSpec,
    row_index: int,
    values: tuple[Any, ...],
) -> tuple[ParsedOrder, list[ImportIssue]]:
    warnings: list[ImportIssue] = []

    raw_status = clean_text(cell(values, spec.status))
    status = map_status(raw_status)
    if status is None:
        status = OrderStatus.SCHEDULE_CONFIRMED
        warnings.append(ImportIssue(sheet, row_index, "warning", f"unknown_status:{raw_status}"))

    received_date = coerce_date(cell(values, spec.received_date))
    scheduled_date = coerce_date(cell(values, spec.scheduled_date))
    if received_date is None:
        received_date = scheduled_date or date.today()
        warnings.append(ImportIssue(sheet, row_index, "warning", "received_date_fallback"))

    schedule_note = clean_text(cell(values, spec.scheduled_date)) if scheduled_date is None else None
    billing_basis = clean_text(cell(values, spec.billing_basis)) if spec.billing_basis else None

    customer_name = single_line(clean_text(cell(values, spec.customer_name)))
    if not customer_name:
        customer_name = single_line(clean_text(cell(values, spec.source_channel))) or "고객"
        warnings.append(ImportIssue(sheet, row_index, "warning", "customer_name_fallback"))
    customer_name = limit_text(customer_name, 80)

    phone_raw = first_present_text(
        cell(values, spec.customer_phone),
        cell(values, spec.team_name),
        cell(values, spec.service_detail),
    )
    customer_phone = first_phone(phone_raw)
    if not customer_phone:
        customer_phone = "0000000000"
        warnings.append(ImportIssue(sheet, row_index, "warning", "customer_phone_fallback"))

    address = clean_text(cell(values, spec.address))
    if not address:
        address = "주소 미입력"
        warnings.append(ImportIssue(sheet, row_index, "warning", "customer_address_fallback"))

    service_name = single_line(clean_text(cell(values, spec.service_name))) or "청소"
    service_name = limit_text(service_name, 160)

    detail_parts = [
        clean_text(cell(values, spec.service_detail)),
        f"방문주기: {schedule_note}" if schedule_note else None,
        f"정산기준: {billing_basis}" if billing_basis else None,
    ]
    service_detail = join_text(*detail_parts)

    source_channel = clean_text(cell(values, spec.source_channel))
    if not source_channel and sheet in {"유니클린텍", "정기청소"}:
        source_channel = sheet

    payment_raw = clean_text(cell(values, spec.payment_status)) if spec.payment_status else None
    payment_status, payment_note = map_customer_payment_status(payment_raw)
    partner_raw = (
        clean_text(cell(values, spec.partner_payment_status))
        if spec.partner_payment_status
        else None
    )
    partner_amount = money_or_none(cell(values, spec.partner_payment_amount))
    partner_status, partner_note, settled_at = map_partner_payment_status(
        partner_raw,
        amount=partner_amount,
        received_date=received_date,
        scheduled_date=scheduled_date,
    )
    if should_gross_up_partner_vat_amount(
        scheduled_date=scheduled_date,
        partner_payment_status=partner_status,
        partner_settled_at=settled_at,
    ):
        partner_amount = gross_up_partner_vat_amount(partner_amount)

    payment_memo = join_text(
        clean_text(cell(values, spec.payment_memo)) if spec.payment_memo else None,
        payment_note,
        partner_note,
    )

    partner_name = first_present_text(
        cell(values, spec.partner_name) if spec.partner_name else None,
        cell(values, spec.team_name),
    )
    partner_name = strip_partner_name(partner_name)

    sheet_slug = spec.slug
    order_id = make_order_id(sheet_slug, row_index)
    return (
        ParsedOrder(
            sheet=sheet,
            sheet_slug=sheet_slug,
            row_index=row_index,
            order_id=order_id,
            preferred_group_id=f"{order_id}-g",
            status=status,
            received_date=received_date,
            scheduled_date=scheduled_date,
            requested_time=limit_text(normalize_time(cell(values, spec.requested_time)), 80),
            team_name=limit_text(single_line(clean_text(cell(values, spec.team_name))), 120),
            service_name=service_name,
            size_or_quantity=limit_text(single_line(clean_text(cell(values, spec.size_or_quantity))), 80),
            service_detail=service_detail,
            special_request=service_detail,
            source_channel=limit_text(single_line(source_channel), 120),
            customer_name=customer_name,
            customer_phone=customer_phone,
            customer_address=address,
            total_amount=money_or_none(cell(values, spec.total_amount)),
            discount_amount=Decimal("0"),
            deposit_amount=money_or_none(cell(values, spec.deposit_amount)),
            balance_amount=money_or_none(cell(values, spec.balance_amount)),
            onsite_extra_amount=money_or_none(cell(values, spec.onsite_extra_amount)),
            vat_type=map_vat_type(clean_text(cell(values, spec.vat_type))),
            payment_status=payment_status,
            payment_memo=payment_memo,
            evidence_memo=clean_text(cell(values, spec.evidence_memo)) if spec.evidence_memo else None,
            partner_name=partner_name,
            partner_payment_amount=partner_amount,
            partner_payment_status=partner_status,
            partner_settled_at=settled_at,
        ),
        warnings,
    )


def apply_rows(db: Session, rows: list[ParsedOrder]) -> ApplyStats:
    ensure_default_partner_categories(db)
    db.flush()
    partner_lookup = build_partner_lookup(db)
    actor_user_id = find_admin_user_id(db)
    business_tz = ZoneInfo(settings.business_timezone)
    stats = ApplyStats()

    for row in rows:
        partner_id, partner_created = ensure_partner(db, row, partner_lookup=partner_lookup)
        stats.partners_created += int(partner_created)
        if partner_created:
            db.flush()

        existing_order = db.get(Order, row.order_id)
        group_id = existing_order.group_id if existing_order and existing_order.group_id else row.preferred_group_id
        group = db.get(OrderGroup, group_id)
        group_created = False
        if group is None:
            group = OrderGroup(
                id=group_id,
                customer_token=token_urlsafe(24),
                customer_name=row.customer_name,
                customer_phone=row.customer_phone,
                customer_address=row.customer_address,
                customer_address_detail=None,
                source_channel=row.source_channel,
                customer_visible_payment=row.customer_visible_payment,
                notes=legacy_note(row),
            )
            db.add(group)
            stats.groups_created += 1
            group_created = True
        else:
            stats.groups_updated += 1

        apply_group_fields(group, row)
        if group_created:
            db.flush()

        if existing_order is None:
            order = Order(id=row.order_id, group_id=group.id, customer_token=group.customer_token)
            db.add(order)
            stats.orders_created += 1
            order_created = True
        else:
            order = existing_order
            stats.orders_updated += 1
            order_created = False

        apply_order_fields(order, row, group=group, partner_id=partner_id)
        if order_created:
            db.flush()

        stats.timeline_events_created += int(
            add_timeline_if_missing(
                db,
                event_id=f"{row.order_id}-created",
                order_id=row.order_id,
                actor_user_id=actor_user_id,
                event_type=TimelineEventType.CREATED,
                title="Legacy order imported",
                description=f"Imported from {row.sheet} row {row.row_index}.",
                metadata=timeline_metadata(row),
            )
        )
        if partner_id:
            stats.timeline_events_created += int(
                add_timeline_if_missing(
                    db,
                    event_id=f"{row.order_id}-partner",
                    order_id=row.order_id,
                    actor_user_id=actor_user_id,
                    event_type=TimelineEventType.PARTNER_ASSIGNED,
                    title="Legacy partner assigned",
                    description=f"Assigned from legacy partner name: {row.partner_name}.",
                    metadata={**timeline_metadata(row), "partner_id": partner_id},
                )
            )
        if row.partner_payment_status == PartnerPaymentStatus.PAID.value:
            stats.timeline_events_created += int(
                add_timeline_if_missing(
                    db,
                    event_id=f"{row.order_id}-settled",
                    order_id=row.order_id,
                    actor_user_id=actor_user_id,
                    event_type=TimelineEventType.PARTNER_SETTLED,
                    title="Legacy partner settlement imported",
                    description="Partner settlement status was imported as paid.",
                    metadata={
                        **timeline_metadata(row),
                        "settled_at": (
                            row.partner_settled_at.astimezone(business_tz).isoformat()
                            if row.partner_settled_at
                            else None
                        ),
                    },
                )
            )

    return stats


def apply_group_fields(group: OrderGroup, row: ParsedOrder) -> None:
    group.customer_name = row.customer_name
    group.customer_phone = row.customer_phone
    group.customer_address = row.customer_address
    group.customer_address_detail = None
    group.source_channel = row.source_channel
    group.customer_visible_payment = row.customer_visible_payment
    group.notes = merge_legacy_note(group.notes, legacy_note(row))


def apply_order_fields(
    order: Order,
    row: ParsedOrder,
    *,
    group: OrderGroup,
    partner_id: str | None,
) -> None:
    order.group_id = group.id
    order.status = row.status
    order.received_date = row.received_date
    order.scheduled_date = row.scheduled_date
    order.requested_time = row.requested_time
    order.partner_id = partner_id
    order.team_name = row.team_name or row.partner_name
    order.service_category_id = None
    order.service_item_id = None
    order.service_name = row.service_name
    order.size_or_quantity = row.size_or_quantity
    order.service_detail = row.service_detail
    order.special_request = row.special_request
    order.source_channel = row.source_channel
    order.customer_name = group.customer_name
    order.customer_phone = group.customer_phone
    order.customer_address = group.customer_address
    order.total_amount = row.total_amount
    order.discount_amount = row.discount_amount
    order.deposit_amount = row.deposit_amount
    order.balance_amount = row.balance_amount
    order.onsite_extra_amount = row.onsite_extra_amount
    order.vat_type = row.vat_type
    order.payment_status = row.payment_status
    order.payment_memo = row.payment_memo
    order.evidence_memo = row.evidence_memo
    order.partner_payment_amount = row.partner_payment_amount
    order.partner_payment_status = row.partner_payment_status
    order.partner_settled_at = row.partner_settled_at
    order.customer_token = group.customer_token
    order.customer_visible_payment = group.customer_visible_payment


def ensure_default_partner_categories(db: Session) -> None:
    for default_category in DEFAULT_PARTNER_CATEGORIES:
        if db.get(PartnerCategory, default_category.id) is None:
            db.add(
                PartnerCategory(
                    id=default_category.id,
                    name=default_category.name,
                    description=default_category.description,
                    is_active=True,
                    sort_order=default_category.sort_order,
                )
            )


def ensure_partner(
    db: Session,
    row: ParsedOrder,
    *,
    partner_lookup: dict[str, Partner],
) -> tuple[str | None, bool]:
    if not row.partner_name:
        return None, False

    key = normalize_partner_key(row.partner_name)
    existing = partner_lookup.get(key)
    if existing is not None:
        return existing.id, False

    partner_id = f"cleanjob-partner-{stable_hash(row.partner_name, 12)}"
    existing_by_id = db.get(Partner, partner_id)
    if existing_by_id is not None:
        partner_lookup[key] = existing_by_id
        return existing_by_id.id, False

    partner = Partner(
        id=partner_id,
        partner_category_id=infer_partner_category_id(row.service_name),
        name=limit_text(row.partner_name, 120) or "Legacy Partner",
        manager_name=None,
        phone=first_phone(row.team_name) or "0000000000",
        service_areas=None,
        available_services=row.service_name,
        memo=f"legacy_import={IMPORT_ID}; source={row.sheet}:{row.row_index}",
        is_active=True,
    )
    db.add(partner)
    partner_lookup[key] = partner
    return partner.id, True


def add_timeline_if_missing(
    db: Session,
    *,
    event_id: str,
    order_id: str,
    actor_user_id: str | None,
    event_type: TimelineEventType,
    title: str,
    description: str,
    metadata: dict[str, Any],
) -> bool:
    if db.get(OrderTimeline, event_id) is not None:
        return False
    db.add(
        OrderTimeline(
            id=event_id,
            order_id=order_id,
            actor_user_id=actor_user_id,
            event_type=event_type,
            title=title,
            description=description,
            event_metadata=metadata,
        )
    )
    return True


def build_database_plan(db: Session, rows: list[ParsedOrder]) -> DatabasePlan:
    order_ids = [row.order_id for row in rows]
    existing_order_ids = set(db.scalars(select(Order.id).where(Order.id.in_(order_ids))))
    partner_lookup = build_partner_lookup(db)
    existing_partner_ids = set(db.scalars(select(Partner.id)))
    existing_partner_keys = set(partner_lookup)
    missing_partner_keys: set[str] = set()
    for row in rows:
        if not row.partner_name:
            continue
        key = normalize_partner_key(row.partner_name)
        partner_id = f"cleanjob-partner-{stable_hash(row.partner_name, 12)}"
        if key and key not in existing_partner_keys and partner_id not in existing_partner_ids:
            missing_partner_keys.add(key)
    return DatabasePlan(
        orders_to_create=len(order_ids) - len(existing_order_ids),
        orders_to_update=len(existing_order_ids),
        partners_to_create=len(missing_partner_keys),
    )


def build_partner_lookup(db: Session) -> dict[str, Partner]:
    lookup: dict[str, Partner] = {}
    for partner in db.scalars(select(Partner)):
        key = normalize_partner_key(partner.name)
        if key and key not in lookup:
            lookup[key] = partner
    return lookup


def find_admin_user_id(db: Session) -> str | None:
    return db.scalar(
        select(User.id)
        .where(User.role == UserRole.ADMIN, User.is_active.is_(True))
        .order_by(User.created_at.asc())
        .limit(1)
    )


def print_summary(*, path: Path, result: ParseResult, plan: DatabasePlan) -> None:
    sheet_counts = Counter(row.sheet for row in result.rows)
    status_counts = Counter(row.status.value for row in result.rows)
    payment_counts = Counter(row.payment_status or "(blank)" for row in result.rows)
    partner_payment_counts = Counter(row.partner_payment_status or "(blank)" for row in result.rows)
    scheduled_dates = [row.scheduled_date for row in result.rows if row.scheduled_date]
    received_dates = [row.received_date for row in result.rows]

    print(f"workbook={path}")
    print(
        f"parsed_rows={len(result.rows)}, skipped_rows={result.skipped_rows}, "
        f"errors={len(result.errors)}, warnings={len(result.warnings)}"
    )
    print(
        "db_plan "
        f"orders_to_create={plan.orders_to_create}, "
        f"orders_to_update={plan.orders_to_update}, "
        f"partners_to_create={plan.partners_to_create}"
    )
    if received_dates:
        print(f"received_date_range={min(received_dates)}..{max(received_dates)}")
    if scheduled_dates:
        print(f"scheduled_date_range={min(scheduled_dates)}..{max(scheduled_dates)}")
    print("sheets=" + format_counter(sheet_counts))
    print("statuses=" + format_counter(status_counts))
    print("payment_statuses=" + format_counter(payment_counts))
    print("partner_payment_statuses=" + format_counter(partner_payment_counts))

    for issue in result.errors[:20]:
        print(f"ERROR {issue.sheet}:{issue.row_index} {issue.reason}")
    for issue in result.warnings[:20]:
        print(f"WARN {issue.sheet}:{issue.row_index} {issue.reason}")
    if len(result.errors) > 20:
        print(f"... {len(result.errors) - 20} more errors")
    if len(result.warnings) > 20:
        print(f"... {len(result.warnings) - 20} more warnings")


def format_counter(counter: Counter[str]) -> str:
    return ", ".join(f"{key}:{value}" for key, value in counter.most_common())


def is_blank_legacy_row(values: tuple[Any, ...], spec: SheetSpec) -> bool:
    meaningful_indexes = [
        spec.received_date,
        spec.scheduled_date,
        spec.team_name,
        spec.address,
        spec.service_detail,
        spec.customer_name,
        spec.customer_phone,
        spec.total_amount,
    ]
    return all(not clean_text(cell(values, index)) for index in meaningful_indexes)


def map_status(raw_status: str | None) -> OrderStatus | None:
    if not raw_status:
        return OrderStatus.SCHEDULE_CONFIRMED
    if raw_status in DIRECT_STATUS_MAP:
        return DIRECT_STATUS_MAP[raw_status]
    if raw_status in LEGACY_STATUS_MAP:
        return LEGACY_STATUS_MAP[raw_status]
    if "취소" in raw_status:
        return OrderStatus.CANCELLED
    if "완료" in raw_status and "계약" not in raw_status and "예약" not in raw_status:
        return OrderStatus.COMPLETED
    return None


def map_customer_payment_status(raw_value: str | None) -> tuple[str | None, str | None]:
    if not raw_value:
        return None, None

    compact = raw_value.replace(" ", "")
    if "환불" in compact:
        return PaymentStatus.REFUNDED.value, None
    if any(token in compact for token in ("최종결제완료", "총액결제완료", "잔금결제완료", "결제완료")):
        return PaymentStatus.PAID.value, None
    if "계약금" in compact:
        return PaymentStatus.DEPOSIT_PAID.value, None
    if any(token in compact for token in ("미입금", "미수", "대기")):
        return PaymentStatus.BALANCE_PENDING.value, None

    return PaymentStatus.PENDING.value, f"결제확인 원문: {raw_value}"


def map_partner_payment_status(
    raw_value: str | None,
    *,
    amount: Decimal | None,
    received_date: date,
    scheduled_date: date | None,
) -> tuple[str | None, str | None, datetime | None]:
    if not raw_value:
        if amount and amount > 0:
            return PartnerPaymentStatus.UNPAID.value, None, None
        return None, None, None

    compact = raw_value.replace(" ", "")
    if any(token in compact for token in ("결제완료", "입금완료", "정산완료")):
        return (
            PartnerPaymentStatus.PAID.value,
            f"협력사 결제확인 원문: {raw_value}",
            parse_settled_at(raw_value, received_date=received_date, scheduled_date=scheduled_date),
        )
    if any(token in compact for token in ("보류", "확인필요", "홀드")):
        return PartnerPaymentStatus.HOLD.value, f"협력사 결제확인 원문: {raw_value}", None

    return PartnerPaymentStatus.READY.value, f"협력사 결제확인 원문: {raw_value}", None


def map_vat_type(raw_value: str | None) -> str | None:
    if not raw_value:
        return None
    if "별도" in raw_value or "미포함" in raw_value:
        return VatType.EXCLUDED.value
    if "포함" in raw_value:
        return VatType.INCLUDED.value
    return None


def parse_settled_at(
    raw_value: str,
    *,
    received_date: date,
    scheduled_date: date | None,
) -> datetime | None:
    match = SETTLED_DATE_PATTERN.search(raw_value)
    if not match:
        return None
    month = int(match.group(1))
    day = int(match.group(2))
    year = (scheduled_date or received_date).year
    try:
        return datetime(year, month, day, tzinfo=ZoneInfo(settings.business_timezone))
    except ValueError:
        return None


def make_order_id(sheet_slug: str, row_index: int) -> str:
    if sheet_slug == "svc":
        return f"cleanjob-{row_index:04d}"
    return f"cleanjob-{sheet_slug}-{row_index:04d}"


def legacy_note(row: ParsedOrder) -> str:
    return f"legacy_import={IMPORT_ID}; source_sheet={row.sheet}; source_row={row.row_index}"


def merge_legacy_note(existing: str | None, note: str) -> str:
    if not existing:
        return note
    if IMPORT_ID in existing:
        lines = [line for line in existing.splitlines() if IMPORT_ID not in line]
        return join_text(*lines, note) or note
    return join_text(existing, note) or note


def timeline_metadata(row: ParsedOrder) -> dict[str, Any]:
    return {
        "source": IMPORT_ID,
        "source_sheet": row.sheet,
        "source_row": row.row_index,
        "legacy_order_id": row.order_id,
    }


def normalize_time(value: Any) -> str | None:
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    return single_line(clean_text(value))


def coerce_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        year = value.year - 2300 if value.year > 3000 else value.year
        return date(year, value.month, value.day)
    if isinstance(value, date):
        year = value.year - 2300 if value.year > 3000 else value.year
        return date(year, value.month, value.day)
    text = clean_text(value)
    if not text:
        return None
    normalized = text.replace(".", "-").replace("/", "-")
    for fmt in ("%Y-%m-%d", "%y-%m-%d"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            pass
    return None


def money_or_none(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    text = clean_text(value)
    if not text:
        return None
    digits = MONEY_PATTERN.sub("", text.replace(",", ""))
    if not digits or digits in {"-", ".", "-."}:
        return None
    try:
        return Decimal(digits)
    except (InvalidOperation, ValueError):
        return None


def first_phone(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    match = PHONE_PATTERN.search(text)
    if match:
        return normalize_phone(match.group(0))
    digits = normalize_phone(text)
    if digits and len(digits) in {10, 11}:
        return digits
    return None


def strip_partner_name(value: Any) -> str | None:
    text = single_line(clean_text(value))
    if not text:
        return None
    text = PHONE_PATTERN.sub("", text)
    text = re.sub(r"\s+", " ", text).strip(" ,-/_")
    return limit_text(text, 120)


def normalize_partner_key(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", "", value).lower()


def first_present_text(*values: Any) -> str | None:
    for value in values:
        text = clean_text(value)
        if text:
            return text
    return None


def join_text(*parts: str | None) -> str | None:
    cleaned = [part.strip() for part in parts if part and part.strip()]
    if not cleaned:
        return None
    return "\n".join(cleaned)


def single_line(value: str | None) -> str | None:
    if not value:
        return None
    return re.sub(r"\s+", " ", value).strip()


def limit_text(value: str | None, max_length: int) -> str | None:
    if value is None or len(value) <= max_length:
        return value
    return value[:max_length]


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return None
    if text in {"-", "--", "ㅡ", "None", "none", "NULL", "null"}:
        return None
    return text


def cell(values: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(values):
        return None
    return values[index]


def stable_hash(value: str, length: int) -> str:
    return hashlib.sha1(value.encode("utf-8")).hexdigest()[:length]


def parse_date_arg(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


if __name__ == "__main__":
    main()
