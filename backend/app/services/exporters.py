from __future__ import annotations

import csv
import io
from collections.abc import Iterable, Mapping, Sequence
from datetime import date
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def to_csv_bytes(headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_to_cell(value) for value in row])
    return buffer.getvalue().encode("utf-8-sig")


def to_xlsx_bytes(
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    sheet_name: str = "report",
    polished: bool = False,
    number_headers: Sequence[str] = (),
    wrap_headers: Sequence[str] = (),
    width_overrides: Mapping[str, float] | None = None,
) -> bytes:
    wb = Workbook(write_only=False)
    try:
        ws = wb.active
        ws.title = sheet_name[:31]
        header_values = list(headers)
        body_rows: list[list[Any]] = []
        ws.append(header_values)
        for row in rows:
            body_row = [_to_cell(value) for value in row]
            body_rows.append(body_row)
            ws.append(body_row)
        if polished:
            _polish_sheet(
                ws,
                header_values,
                body_rows,
                number_headers=set(number_headers),
                wrap_headers=set(wrap_headers),
                width_overrides=width_overrides or {},
            )
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    finally:
        wb.close()


def _to_cell(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, date):
        return value.isoformat()
    return value


def _polish_sheet(
    ws,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    number_headers: set[str],
    wrap_headers: set[str],
    width_overrides: Mapping[str, float],
) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    thin_side = Side(style="thin", color="E5E7EB")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = border
    ws.row_dimensions[1].height = 24

    for column_index, header in enumerate(headers, start=1):
        letter = get_column_letter(column_index)
        values = [header, *[row[column_index - 1] if column_index <= len(row) else "" for row in rows]]
        ws.column_dimensions[letter].width = width_overrides.get(header, _suggest_width(values, header in wrap_headers))

        for row_index in range(2, len(rows) + 2):
            cell = ws.cell(row=row_index, column=column_index)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if header in number_headers else "left",
                vertical="top",
                wrap_text=header in wrap_headers,
            )
            if header in number_headers:
                cell.number_format = "#,##0"


def _suggest_width(values: Sequence[Any], should_wrap: bool) -> float:
    max_width = max((_display_width(value) for value in values), default=0)
    cap = 42 if should_wrap else 24
    floor = 12 if should_wrap else 9
    return float(min(max(max_width + 2, floor), cap))


def _display_width(value: Any) -> int:
    if value is None:
        return 0
    return max(
        (
            sum(2 if ord(char) > 127 else 1 for char in line)
            for line in str(value).splitlines() or [""]
        ),
        default=0,
    )
