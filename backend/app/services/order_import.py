from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.time import business_today
from app.domain.constants import OrderStatus
from app.domain.phone import normalize_phone
from app.schemas.order import OrderGroupCreate, OrderLineCreate
from app.schemas.report import OrderImportFailure, OrderImportResult
from app.services.orders import OrderService

REQUIRED_HEADERS = [
    "group_key",
    "customer_name",
    "customer_phone",
    "customer_address",
    "scheduled_date",
    "service_name",
    "total_amount",
]

_XLSX_MEDIA_PREFIXES = (
    "application/vnd.openxmlformats",
    "application/vnd.ms-excel",
    "application/octet-stream",
)


def is_xlsx_upload(filename: str | None, content_type: str | None) -> bool:
    name = (filename or "").lower()
    suffix = Path(name).suffix
    if suffix:
        return suffix == ".xlsx"
    return bool(content_type and content_type.startswith(_XLSX_MEDIA_PREFIXES))


def import_orders_from_xlsx(
    *,
    file_bytes: bytes,
    filename: str | None,
    content_type: str | None,
    db: Session,
    actor_user_id: str,
) -> OrderImportResult:
    if not is_xlsx_upload(filename, content_type):
        raise ValueError(f"unsupported_content_type:{content_type or '(none)'}")

    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"invalid_xlsx:{exc}") from exc

    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = list(next(rows_iter, []) or [])
        header_to_idx = {str(header).strip(): idx for idx, header in enumerate(headers) if header is not None}
        missing = [header for header in REQUIRED_HEADERS if header not in header_to_idx]
        if missing:
            return OrderImportResult(
                succeeded_groups=0,
                succeeded_lines=0,
                failed=[
                    OrderImportFailure(
                        row_index=0,
                        reason=f"missing_columns:{','.join(missing)}",
                    )
                ],
            )

        grouped: dict[str, dict[str, Any]] = {}
        group_errors: dict[str, list[OrderImportFailure]] = {}

        for row_index, raw in enumerate(rows_iter, start=1):
            if raw is None or _is_empty_row(raw):
                continue

            group_key_raw = raw[header_to_idx["group_key"]]
            group_key = str(group_key_raw or "").strip()
            try:
                line = _parse_row(raw, header_to_idx)
            except ValueError as exc:
                key = group_key if group_key else f"__unknown_row_{row_index}"
                group_errors.setdefault(key, []).append(
                    OrderImportFailure(row_index=row_index, reason=str(exc))
                )
                continue

            if group_key not in grouped:
                grouped[group_key] = {
                    "customer_name": line["customer_name"],
                    "customer_phone": line["customer_phone"],
                    "customer_address": line["customer_address"],
                    "lines": [],
                }
            else:
                head = grouped[group_key]
                if (
                    head["customer_name"] != line["customer_name"]
                    or head["customer_phone"] != line["customer_phone"]
                    or head["customer_address"] != line["customer_address"]
                ):
                    group_errors.setdefault(group_key, []).append(
                        OrderImportFailure(
                            row_index=row_index,
                            reason="inconsistent_group_customer",
                        )
                    )
                    continue

            grouped[group_key]["lines"].append({"row_index": row_index, "data": line})
    finally:
        wb.close()

    failed: list[OrderImportFailure] = []
    for group_key, errors in group_errors.items():
        if group_key.startswith("__unknown_row_"):
            failed.extend(errors)
            continue
        failed.extend(errors)
        error_indices = {error.row_index for error in errors}
        ok_rows = [entry["row_index"] for entry in grouped.get(group_key, {}).get("lines", [])]
        for row_index in ok_rows:
            if row_index not in error_indices:
                failed.append(
                    OrderImportFailure(
                        row_index=row_index,
                        reason="group_skipped_due_to_sibling_error",
                    )
                )
        grouped.pop(group_key, None)

    service = OrderService(db)
    succeeded_groups = 0
    succeeded_lines = 0
    received_date = business_today()

    for bundle in grouped.values():
        line_payloads = [
            OrderLineCreate(
                status=OrderStatus.NEW,
                received_date=received_date,
                scheduled_date=entry["data"]["scheduled_date"],
                service_name=entry["data"]["service_name"],
                total_amount=float(entry["data"]["total_amount"]),
            )
            for entry in bundle["lines"]
        ]
        try:
            service.create_group(
                OrderGroupCreate(
                    customer_name=bundle["customer_name"],
                    customer_phone=bundle["customer_phone"],
                    customer_address=bundle["customer_address"],
                    notes=None,
                    lines=line_payloads,
                ),
                actor_user_id=actor_user_id,
            )
            succeeded_groups += 1
            succeeded_lines += len(line_payloads)
        except Exception as exc:  # noqa: BLE001
            db.rollback()
            for entry in bundle["lines"]:
                failed.append(
                    OrderImportFailure(
                        row_index=entry["row_index"],
                        reason=f"group_create_failed:{exc}",
                    )
                )

    failed.sort(key=lambda failure: failure.row_index)
    return OrderImportResult(
        succeeded_groups=succeeded_groups,
        succeeded_lines=succeeded_lines,
        failed=failed,
    )


def _is_empty_row(raw: tuple[Any, ...]) -> bool:
    return all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in raw)


def _parse_row(raw: tuple[Any, ...], header_to_idx: dict[str, int]) -> dict[str, Any]:
    def cell(name: str) -> Any:
        return raw[header_to_idx[name]]

    group_key = str(cell("group_key") or "").strip()
    if not group_key:
        raise ValueError("group_key_required")

    customer_name = str(cell("customer_name") or "").strip()
    if not customer_name:
        raise ValueError("customer_name_required")

    phone_raw = str(cell("customer_phone") or "").strip()
    phone = normalize_phone(phone_raw)
    if not phone or not phone.isdigit() or len(phone) not in {10, 11}:
        raise ValueError(f"invalid_phone:{phone_raw}")

    customer_address = str(cell("customer_address") or "").strip()
    if not customer_address:
        raise ValueError("customer_address_required")

    service_name = str(cell("service_name") or "").strip()
    if not service_name:
        raise ValueError("service_name_required")

    total_amount = _coerce_decimal(cell("total_amount"))
    if total_amount < 0:
        raise ValueError("negative_total_amount")

    return {
        "group_key": group_key,
        "customer_name": customer_name,
        "customer_phone": phone,
        "customer_address": customer_address,
        "scheduled_date": _coerce_date(cell("scheduled_date")),
        "service_name": service_name,
        "total_amount": total_amount,
    }


def _coerce_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError as exc:
            raise ValueError(f"invalid_scheduled_date:{value}") from exc
    raise ValueError("invalid_scheduled_date")


def _coerce_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"invalid_total_amount:{value}") from exc
